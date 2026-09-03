"""Tests for stakeholder reporting (reporting.py).

Run with:  python -m unittest project_management.tests.test_reporting -v
"""

import json
import tempfile
import unittest
from datetime import date
from pathlib import Path

from ..reporting import (
    build_report,
    collect_rows,
    diff_reports,
    parse_questions,
    save_snapshot,
    snapshot_for_report,
    write_report_xlsx,
)


def _write_draft(tmp, name, content):
    path = Path(tmp) / name
    path.write_text(content, encoding="utf-8")
    return path


def _seed_drafts(directory):
    """Three drafts: 01 stamped+closed, 02 stamped+open (depends on 01), 03 planned."""
    _write_draft(
        directory,
        "01-contract.md",
        "# Task: Contract\n"
        "## Metadata\n"
        "- **Deadline:** 2026-09-04\n"
        "- **Status:** TODO\n"
        "- **Assignee:** Team\n"
        "- **Tags:** gate:1-decisions\n"
        "- **Dependencies:** None\n"
        "- **GitHub Issue:** #1 (https://github.com/x/y/issues/1)\n"
        "- **Estimated Effort:** 2h\n"
        "\n## Questions to Resolve\n"
        "- [ ] Which framework?\n",
    )
    _write_draft(
        directory,
        "02-db-choice.md",
        "# Task: DB choice\n"
        "## Metadata\n"
        "- **Deadline:** 2026-09-04\n"
        "- **Status:** TODO\n"
        "- **Assignee:** Person 1\n"
        "- **Tags:** gate:1-decisions\n"
        "- **Dependencies:** 01-contract.md\n"
        "- **GitHub Issue:** #2 (https://github.com/x/y/issues/2)\n"
        "- **Estimated Effort:** 1h\n"
        "\n## Questions to Resolve\n"
        "- [x] SQLite or PostgreSQL?\n"
        "- [ ] Docker needed?\n",
    )
    _write_draft(
        directory,
        "03-artifacts.md",
        "# Task: Artifacts\n"
        "## Metadata\n"
        "- **Deadline:** 2026-09-05\n"
        "- **Status:** TODO\n"
        "- **Assignee:** TBD\n"
        "- **Tags:** gate:1-decisions, stream:d-audit\n"
        "- **Dependencies:** None\n"
        "- **Estimated Effort:** 2h\n",
    )


class TestParseQuestions(unittest.TestCase):
    def test_parses_open_and_resolved(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = _write_draft(
                tmp,
                "01-a.md",
                "# Task: A\n## Questions to Resolve\n- [ ] Open one\n- [x] Resolved one\n- plain note\n",
            )
            questions = parse_questions(path)
        self.assertEqual(len(questions), 3)
        self.assertFalse(questions[0]["resolved"])
        self.assertTrue(questions[1]["resolved"])
        self.assertFalse(questions[2]["resolved"])

    def test_no_section_returns_empty(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = _write_draft(tmp, "01-b.md", "# Task: B\n## Requirements\n- stuff\n")
            self.assertEqual(parse_questions(path), [])


class TestCollectRows(unittest.TestCase):
    def test_merges_github_state(self):
        with tempfile.TemporaryDirectory() as tmp:
            _seed_drafts(tmp)
            github = {
                1: {"state": "closed", "url": "https://github.com/x/y/issues/1"},
                2: {"state": "open", "url": "https://github.com/x/y/issues/2"},
            }
            rows = collect_rows(tmp, github, today=date(2026, 9, 10))
        by_key = {row["key"]: row for row in rows}
        self.assertEqual(len(rows), 3)
        self.assertEqual(by_key["01"]["state"], "Closed")
        self.assertEqual(by_key["02"]["state"], "Open")
        self.assertEqual(by_key["02"]["blocked_by"], "#1")
        self.assertTrue(by_key["02"]["blocked"])
        self.assertFalse(by_key["01"]["blocked"])
        self.assertEqual(by_key["03"]["state"], "Planned")
        self.assertEqual(by_key["03"]["stream"], "D-Audit")
        self.assertTrue(by_key["03"]["unassigned"])
        self.assertTrue(by_key["01"]["unassigned"])  # assignee Team
        self.assertFalse(by_key["02"]["unassigned"])
        self.assertEqual(len(by_key["01"]["questions"]), 1)
        self.assertEqual(len(by_key["02"]["questions"]), 2)

    def test_overdue_flags_only_non_closed(self):
        with tempfile.TemporaryDirectory() as tmp:
            _seed_drafts(tmp)
            github = {1: {"state": "closed", "url": ""}, 2: {"state": "open", "url": ""}}
            rows = collect_rows(tmp, github, today=date(2026, 9, 10))
        by_key = {row["key"]: row for row in rows}
        # 02 open with deadline 2026-09-04 before today -> overdue
        self.assertTrue(by_key["02"]["overdue"])
        # 01 closed with past deadline -> not overdue
        self.assertFalse(by_key["01"]["overdue"])


class TestSnapshotAndChangelog(unittest.TestCase):
    def _two_reports(self, tmp):
        """Build report 1 from fixture drafts, then modify task 02 and re-build."""
        _seed_drafts(tmp)
        github = {
            1: {"state": "closed", "url": "https://github.com/x/y/issues/1"},
            2: {"state": "open", "url": "https://github.com/x/y/issues/2"},
        }
        first = build_report(tmp, github, today=date(2026, 9, 10))
        # Task 02 changes: now closed on GitHub, deadline moved, reassigned,
        # question resolved.
        (Path(tmp) / "02-db-choice.md").write_text(
            "# Task: DB choice\n"
            "## Metadata\n"
            "- **Deadline:** 2026-09-06\n"
            "- **Status:** TODO\n"
            "- **Assignee:** Person 2\n"
            "- **Tags:** gate:1-decisions\n"
            "- **Dependencies:** 01-contract.md\n"
            "- **GitHub Issue:** #2 (https://github.com/x/y/issues/2)\n"
            "- **Estimated Effort:** 1h\n"
            "\n## Questions to Resolve\n"
            "- [x] SQLite or PostgreSQL?\n"
            "- [x] Docker needed?\n",
            encoding="utf-8",
        )
        github[2] = {"state": "closed", "url": "https://github.com/x/y/issues/2"}
        # Task 03 gets a GitHub card between the two exports.
        (Path(tmp) / "03-artifacts.md").write_text(
            "# Task: Artifacts\n"
            "## Metadata\n"
            "- **Deadline:** 2026-09-05\n"
            "- **Status:** TODO\n"
            "- **Assignee:** Person 3\n"
            "- **Tags:** gate:1-decisions, stream:d-audit\n"
            "- **Dependencies:** None\n"
            "- **GitHub Issue:** #3 (https://github.com/x/y/issues/3)\n"
            "- **Estimated Effort:** 2h\n",
            encoding="utf-8",
        )
        github[3] = {"state": "open", "url": "https://github.com/x/y/issues/3"}
        second = build_report(tmp, github, today=date(2026, 9, 10))
        return first, second

    def test_no_previous_snapshot_means_no_entries(self):
        with tempfile.TemporaryDirectory() as tmp:
            first, _second = self._two_reports(tmp)
            self.assertEqual(diff_reports(None, first), [])

    def test_diff_detects_task_and_question_changes(self):
        with tempfile.TemporaryDirectory() as tmp:
            first, second = self._two_reports(tmp)
            snapshot = snapshot_for_report(first, generated_at="2026-09-10 08:00 UTC")
            entries = diff_reports(snapshot, second)
        task_entries = [entry for entry in entries if entry["sheet"] == "Tasks"]
        changes = {entry["change"] for entry in task_entries if entry["task"].startswith("02")}
        self.assertIn("State", changes)
        self.assertIn("Deadline", changes)
        self.assertIn("Assignee", changes)
        self.assertTrue(
            any(entry["change"] == "GitHub card" for entry in task_entries),
            task_entries,
        )
        question_entries = [entry for entry in entries if entry["sheet"] == "Questions"]
        self.assertTrue(
            any(entry["change"] == "Resolved" for entry in question_entries),
            question_entries,
        )

    def test_snapshot_roundtrip_identical_report_no_entries(self):
        with tempfile.TemporaryDirectory() as tmp:
            first, _second = self._two_reports(tmp)
            snapshot_path = Path(tmp) / "report.snapshot.json"
            save_snapshot(snapshot_path, snapshot_for_report(first))
            loaded = json.loads(snapshot_path.read_text(encoding="utf-8"))
            # Same data in, no changes out
            self.assertEqual(diff_reports(loaded, first), [])


class TestBuildReport(unittest.TestCase):
    def test_summary_and_rollups(self):
        with tempfile.TemporaryDirectory() as tmp:
            _seed_drafts(tmp)
            github = {1: {"state": "closed", "url": ""}, 2: {"state": "open", "url": ""}}
            report = build_report(tmp, github, today=date(2026, 9, 10))
        self.assertEqual(report["summary"]["planned"], 1)
        self.assertEqual(report["summary"]["open"], 1)
        self.assertEqual(report["summary"]["closed"], 1)
        self.assertEqual(report["summary"]["blocked"], 1)
        # 02 is open with a past deadline and 03 is planned with a past deadline
        self.assertEqual(report["summary"]["overdue"], 2)
        self.assertEqual(report["summary"]["unassigned"], 2)
        # One gate with all three tasks
        self.assertEqual(report["gates"][0]["name"], "1-Decisions")
        self.assertEqual(report["gates"][0]["tasks"], 3)
        self.assertEqual(report["gates"][0]["closed"], 1)
        # One stream (D-Audit) plus none for the untagged rows
        stream_names = {stream["name"] for stream in report["streams"]}
        self.assertIn("D-Audit", stream_names)
        # Two unresolved questions exported
        open_questions = [q for q in report["questions"] if not q["resolved"]]
        self.assertEqual(len(open_questions), 2)


class TestWriteXlsx(unittest.TestCase):
    def test_workbook_has_four_sheets(self):
        import openpyxl

        with tempfile.TemporaryDirectory() as tmp:
            _seed_drafts(tmp)
            report = build_report(tmp, {1: {"state": "open", "url": ""}}, today=date(2026, 9, 10))
            output = Path(tmp) / "report.xlsx"
            changelog = {
                "baseline": "2026-09-09 10:00 UTC",
                "entries": [
                    {
                        "sheet": "Tasks",
                        "task": "02 DB choice",
                        "change": "State",
                        "before": "Open",
                        "after": "Closed",
                    }
                ],
            }
            write_report_xlsx(
                str(output),
                report,
                generated_at="2026-09-10",
                repository="x/y",
                changelog=changelog,
            )

            wb = openpyxl.load_workbook(output)
            self.assertEqual(wb.sheetnames, ["Overview", "Changelog", "Tasks", "Questions"])
            changelog_sheet = wb["Changelog"]
            self.assertIn("changes since the previous export", changelog_sheet["A1"].value)
            # Baseline header row 4, change row 5
            self.assertEqual(changelog_sheet.cell(row=5, column=3).value, "State")
            tasks = wb["Tasks"]
            # Header + 3 task rows
            self.assertEqual(tasks.max_row, 4)
            headers = [tasks.cell(row=1, column=i).value for i in range(1, 5)]
            self.assertEqual(headers[:4], ["Key", "Epic (Gate)", "Stream", "Title"])
            overview = wb["Overview"]
            self.assertIn("HealthAccess", overview["A1"].value)
            questions = wb["Questions"]
            texts = [
                str(questions.cell(row=r, column=3).value) for r in range(2, questions.max_row + 1)
            ]
            self.assertIn("Which framework?", texts)


if __name__ == "__main__":
    unittest.main()
